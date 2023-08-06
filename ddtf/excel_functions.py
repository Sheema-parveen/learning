from openpyxl import load_workbook
#methods to read and write functions in excel file
class sheema_excel_function:
    def __init__(self,file_name,sheet_name):
        self.file=file_name
        self.sheet=sheet_name
#count rows in excel file
    def row_count(self):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        return sheet.max_row
    
#count coloumns in excel file
    def column_count(self):
        workbook=load_workbook(self.file)
        sheet=workbook[self.sheet]
        return sheet.max_column
#read data
    def read_data(self,row_number,column_number):
        workbook= load_workbook(self.file)
        sheet=workbook[self.sheet]
        return sheet.cell(row= row_number,column= column_number).value
#write data
    def write_data(self,row_number,column_number,data):
        workbook= load_workbook(self.file)
        sheet=workbook[self.sheet]
        sheet.cell(row= row_number,column= column_number).value = data
        workbook.save(self.file)